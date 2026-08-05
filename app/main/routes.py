# -*- coding: utf-8 -*-
from flask import Blueprint, render_template, jsonify, request
from flask_login import login_required, current_user
from app import db
from app.ai_service import AIService
from app.models import VegetablePrice
from app.utils import sync_missing_data, get_top_rankings, get_vegetable_price_history, get_rankings_with_pandas
from app.utils import calculate_price_change_rate, get_total_count, get_avg_price_today
from datetime import date
from flask import send_file

main_bp = Blueprint('main', __name__)


@main_bp.route('/')
@login_required
def index():
    return render_template('layout.html')


@main_bp.route('/api/overview')
@login_required
def api_overview():
    try:
        from app.models import PriceHistory, VegetablePrice
        from app.utils import get_rankings_with_pandas
        from sqlalchemy import func
        from datetime import date, timedelta

        # 1. 数据统计卡片
        total_count = PriceHistory.query.count()

        # 2. 今日均价
        today = date.today()
        today_prices = VegetablePrice.query.filter_by(source_date=today).all()
        if today_prices:
            avg_price = sum(p.price for p in today_prices) / len(today_prices)
        else:
            avg_price = 0

        # 3. 涨幅跌幅排行
        top_increase, top_decrease = get_rankings_with_pandas(limit=6)

        # 4. 饼图
        highest_prices = db.session.query(
            PriceHistory.prod_name,
            func.max(PriceHistory.price).label('max_price')
        ).group_by(PriceHistory.prod_name).order_by(func.max(PriceHistory.price).desc()).limit(10).all()

        pie_data = []
        for item in highest_prices:
            pie_data.append({
                'name': item[0],
                'value': round(float(item[1]), 2)
            })

        # 5. 近7日整体价格趋势
        start_date_7d = today - timedelta(days=7)
        daily_avg_prices = db.session.query(
            PriceHistory.record_date,
            func.avg(PriceHistory.price).label('avg_price')
        ).filter(
            PriceHistory.record_date >= start_date_7d,
            PriceHistory.price > 0
        ).group_by(PriceHistory.record_date).order_by(PriceHistory.record_date.asc()).all()

        trend_data = []
        for item in daily_avg_prices:
            trend_data.append({
                'date': item[0].strftime('%m/%d'),
                'price': round(float(item[1]), 2)
            })

        return render_template('pages/overview.html',
                               total_count=total_count,
                               avg_price=avg_price,
                               top_increase=top_increase,
                               top_decrease=top_decrease,
                               today=today,
                               pie_data=pie_data,
                               trend_data=trend_data)
    except Exception as e:
        print("Overview error: {}".format(str(e)))
        return jsonify({'error': str(e)}), 500


@main_bp.route('/api/data')
@login_required
def api_data():
    return render_template('pages/data.html')


@main_bp.route('/api/data/list')
@login_required
def api_data_list():
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        vegetable_name = request.args.get('vegetable_name', '', type=str)
        record_date = request.args.get('record_date', '', type=str)

        from app.models import PriceHistory
        query = PriceHistory.query

        if vegetable_name:
            query = query.filter(PriceHistory.prod_name.like('%{}%'.format(vegetable_name)))

        if record_date:
            query = query.filter(PriceHistory.record_date == record_date)

        query = query.order_by(PriceHistory.record_date.desc(), PriceHistory.prod_name)

        pagination = query.paginate(page=page, per_page=per_page, error_out=False)

        data_list = []
        for item in pagination.items:
            data_list.append({
                'id': item.id,
                'vegetable_name': item.prod_name,  # 修改：使用 prod_name
                'category': '蔬菜',
                'price': item.price,
                'min_price': item.min_price,
                'max_price': item.max_price,
                'place': item.place if hasattr(item, 'place') and item.place else '-',
                'unit': item.unit if hasattr(item, 'unit') and item.unit else '斤',
                'record_date': item.record_date.strftime('%Y-%m-%d')
            })

        return jsonify({
            'success': True,
            'data': data_list,
            'total': pagination.total,
            'page': page,
            'per_page': per_page,
            'pages': pagination.pages
        })
    except Exception as e:
        print("Data list error: {}".format(str(e)))
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/api/vegetable_names')
@login_required
def api_vegetable_names():
    try:
        from app.models import PriceHistory
        from sqlalchemy import distinct

        names = db.session.query(distinct(PriceHistory.prod_name)).order_by(PriceHistory.prod_name).all()
        name_list = [n[0] for n in names if n[0]]

        return jsonify({'success': True, 'data': name_list})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/api/export_excel', methods=['GET'])
@login_required
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        from app.models import PriceHistory

        vegetable_name = request.args.get('vegetable_name', '', type=str)
        record_date = request.args.get('record_date', '', type=str)

        query = PriceHistory.query

        if vegetable_name:
            query = query.filter(PriceHistory.prod_name.like('%{}%'.format(vegetable_name)))
        if record_date:
            query = query.filter(PriceHistory.record_date == record_date)

        query = query.order_by(PriceHistory.record_date.desc(), PriceHistory.prod_name)

        data = query.all()

        if not data:
            return jsonify({'success': False, 'message': '没有数据可导出'}), 400

        print("导出数据量: {} 条".format(len(data)))

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "蔬菜价格数据"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2e7d32", end_color="2e7d32", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 写入表头
        headers = ['序号', '蔬菜名称', '价格(元/斤)', '最低价', '最高价', '产地', '单位', '日期']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 写入数据
        for i, item in enumerate(data, 1):
            ws.cell(row=i + 1, column=1, value=i)
            ws.cell(row=i + 1, column=2, value=item.prod_name)
            ws.cell(row=i + 1, column=3, value=float(item.price) if item.price else 0)
            ws.cell(row=i + 1, column=4, value=float(item.min_price) if item.min_price else '-')
            ws.cell(row=i + 1, column=5, value=float(item.max_price) if item.max_price else '-')
            ws.cell(row=i + 1, column=6, value=item.place if item.place else '-')
            ws.cell(row=i + 1, column=7, value=item.unit if item.unit else '斤')
            ws.cell(row=i + 1, column=8, value=item.record_date.strftime('%Y-%m-%d'))

        # 调整列宽
        column_widths = [8, 18, 12, 10, 10, 15, 8, 12]
        for i, width in enumerate(column_widths, 1):
            col_letter = openpyxl.utils.get_column_letter(i)
            ws.column_dimensions[col_letter].width = width

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # 生成文件名
        filename = "vegetable_prices_all_{}.xlsx".format(date.today().strftime('%Y%m%d'))
        if vegetable_name:
            filename = "{}_{}.xlsx".format(vegetable_name, date.today().strftime('%Y%m%d'))
        if record_date:
            filename = "price_{}.xlsx".format(record_date)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print("导出错误: {}".format(str(e)))
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/api/analysis')
@login_required
def api_analysis():
    try:
        from app.models import PriceHistory
        from sqlalchemy import distinct

        # 获取历史数据中所有不同的蔬菜名称
        vegetables = db.session.query(
            distinct(PriceHistory.prod_name).label('name')
        ).order_by(PriceHistory.prod_name).all()

        # 转换为列表格式
        vegetable_list = [{'name': v[0]} for v in vegetables if v[0]]

        return render_template('pages/analysis.html', vegetables=vegetable_list)
    except Exception as e:
        print("Analysis error: {}".format(str(e)))
        return jsonify({'error': str(e)}), 500


@main_bp.route('/api/vegetable_detail/<string:name>')
@login_required
def api_vegetable_detail(name):
    try:
        today = date.today()
        vegetable = VegetablePrice.query.filter_by(name=name, source_date=today).first()

        if not vegetable:
            vegetable = VegetablePrice.query.filter_by(name=name).order_by(VegetablePrice.source_date.desc()).first()

        if not vegetable:
            return jsonify({'error': '蔬菜不存在'}), 404

        history = get_vegetable_price_history(name, days=30)
        yesterday_change = calculate_price_change_rate(vegetable.price, vegetable.yesterday_price)

        return jsonify({
            'name': vegetable.name,
            'category': vegetable.category,
            'price': vegetable.price,
            'yesterday_price': vegetable.yesterday_price,
            'change_rate': yesterday_change,
            'avg_price_7d': vegetable.avg_price_7d,
            'avg_price_20d': vegetable.avg_price_20d,
            'history': history
        })
    except Exception as e:
        print("Detail error: {}".format(str(e)))
        return jsonify({'error': str(e)}), 500


@main_bp.route('/api/sync_data', methods=['POST'])
@login_required
def sync_data():
    """智能同步数据（只同步缺失的日期）"""
    try:
        from app.utils import sync_missing_data
        result = sync_missing_data()
        return jsonify(result)
    except Exception as e:
        print("同步错误: {}".format(str(e)))
        return jsonify({'success': False, 'message': str(e)}), 500

# AI服务实例
ai_service = AIService()


@main_bp.route('/api/ai_chat')
@login_required
def ai_chat_page():
    """AI助手页面"""
    return render_template('pages/ai_chat.html')


@main_bp.route('/api/ai/chat', methods=['POST'])
@login_required
def ai_chat():
    """AI智能问答"""
    data = request.get_json()
    question = data.get('question', '')

    if not question:
        return jsonify({'success': False, 'error': '问题不能为空'}), 400

    answer = ai_service.chat(question)
    return jsonify({'success': True, 'answer': answer})


@main_bp.route('/api/ai/predict/<string:name>', methods=['GET'])
@login_required
def ai_predict(name):
    """价格预测"""
    prediction = ai_service.predict_price(name)
    return jsonify({'success': True, 'prediction': prediction})


@main_bp.route('/api/ai/advice', methods=['GET'])
@login_required
def ai_advice():
    """采购建议"""
    advice = ai_service.get_advice()
    return jsonify({'success': True, 'advice': advice})